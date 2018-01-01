from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from threading import Thread

from commodities import category_list


def write2Excel(trade_date,sheet_name,header_arr,crawl_result):
    path = r"{}.xlsx".format(*trade_date)

    # Check if file is existed
    try:
        book = load_workbook(path)
    except:
        wb = openpyxl.Workbook()
        wb.save(path)
        book = load_workbook(path)

    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = book

    # create header
    df = pd.DataFrame(columns=header_arr)

    # Write loop data
    for i in range(len(crawl_result)):
        df.loc[i] = crawl_result[i]
    df.to_excel(writer,sheet_name=sheet_name, index=False)

    writer.save()
    writer.close()

def FX_get_href(category):
    settlements_url_list = []
    # Parse HTML
    URL = "http://www.cmegroup.com/trading/{}/".format(category)
    html = requests.get(URL)
    soup = BeautifulSoup(html.text, "html.parser")
    data = soup.find("table", id="cmeDelayedQuotes2").find("tbody").findAll("tr")
    FX_href_list = [href.find("a",href=True)["href"] for href in data]

    for FX in FX_href_list:
        URL = "http://www.cmegroup.com{}".format(FX)
        html = requests.get(URL)
        soup = BeautifulSoup(html.text, "html.parser")
        data = soup.find("div", id="productTabs").find("ul",class_="cmeTabsSystem").findAll("li")

        for li in data:
            settlements_url = li.find("a", title="Settlements")
            if settlements_url:
                settlements_url_list.append(settlements_url["href"])

    return settlements_url_list

def CrawlData():
    for category in category_list:
        for url in FX_get_href(category):

            crawl_result = []
            month = ""
            open_price = ""
            high_price = ""
            low = ""
            last = ""
            change = ""
            settle = ""
            estimated_volume = ""
            prior_day_open_interest = ""

            # Parse HTML
            URL = "http://www.cmegroup.com{}".format(url)
            html = requests.get(URL)
            soup = BeautifulSoup(html.text, "html.parser")

            # Product Name
            product_name = soup.find("span", id="productName").text.replace("/"," - ").strip()
            print(product_name)

            # CME trade date
            trade_date = soup.find("select", id="cmeTradeDate").find("option" ,  selected="selected")(text=True)

            # Find table-data
            table = soup.find("table", id="settlementsFuturesProductTable")

            # Dataframe header
            list_header = [h_name.text for h_name in table.find("tr").findAll("th")]
            # Data table
            for row in range(len(table.findAll("tr"))):
                td = table.findAll("tr")[row].findAll("td")
                th = table.findAll("tr")[row].findAll("th")

                # Check number td
                if len(td) == 8:
                    month = th[0].find(text=True)
                    open_price = td[0].find(text=True)
                    high_price = td[1].find(text=True)
                    low = td[2].find(text=True)
                    last = td[3].find(text=True)
                    change = td[4].find(text=True)
                    settle = td[5].find(text=True)
                    estimated_volume = td[6].find(text=True)
                    prior_day_open_interest = td[7].find(text=True)

                    # Append to crawl_result
                    crawl_result.append([month,open_price,high_price,low,last,change,settle,estimated_volume,prior_day_open_interest])

            # Write into xlsx file
            write2Excel(trade_date, product_name.upper(), list_header, crawl_result)

            print (product_name,": Done")


if __name__ == "__main__":
    Thread(target=CrawlData).start()


